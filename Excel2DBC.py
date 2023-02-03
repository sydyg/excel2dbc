#需要根据excel模板配置部分内容，联系作者(sydygys)配置
import openpyxl
import os
import traceback
import sys
#定义信号
class Msg:
    def __init__(self):
        self.Name = ''
        self.ID = 0
        self.PGN = ''
        self.DLC = ''  #有些DLC是Var
        self.TxNode = ''
        self.Description = ''
        self.SendType = ''
        self.CycleTime = ''
        self.Signals = []

class Sig:
    def __init__(self):
        self.Name = ''
        self.Desc = ''
        self.SPN = ''
        self.Byte = ''
        self.Bit = ''
        self.StartBit = 0
        self.BitLenth = ''
        self.SendType = ''
        self.Reso = ''
        self.OffSet = ''
        self.Min = ''
        self.Max = ''
        self.Unit = ''
        self.SigValueDesc = ''
        self.ValueTable = [] #以[value,desc]的形式存储
        self.TxNode = 'Vector__XXX'

#定义各项所在列索引


#获取当前路径
pwd = os.getcwd()
#加载excel文件
def loadexcel(filename):
    wb = openpyxl.load_workbook(filename)
    ws = wb.active #获取表单
    print(ws) #打印表单名称
    print('Excel文件共：{}行 {}列'.format(ws.max_row, ws.max_column))#
    return ws

#获取所有节点
ALLNodes = []
def getAllNodes():
    global ALLNodes
    for i in range(26, MaxColumn):
        ALLNodes.append(ws.cell(row=1, column=i).value)#追加元素到列表
        #print(ALLNodes[i-26])

ExceptionInfo = 'ExEc Error:\n'
#获取Matrix中最原始的信息
def GetMsg():
    CanMsgs = []
    global ExceptionInfo
    getAllNodes()
    Row = 0
    try:
        for Row in range(2, MaxRow+1):#Row = [2:MaxRow]
            MsgName = ws.cell(row=Row,column=MsgName_Index).value
            if(MsgName!=None):   # 报文
                CanMsg = Msg()  # 构造Msg对象
                CanMsg.Name = MsgName
                #print(CanMsg.Name)
                CanMsg.Description = ws.cell(row=Row, column=MsgDesc_Index).value
                CanMsg.ID = int(ws.cell(row=Row, column=MsgID_Index).value,16)+ 0x80000000
                CanMsg.PGN = ws.cell(row=Row, column=MsgPGN_Index).value
                CanMsg.SendType = ws.cell(row=Row, column=MsgSendType_Index).value
                CanMsg.CycleTime = ws.cell(row=Row, column=MsgCycle_Index).value
                CanMsg.DLC = ws.cell(row=Row, column=MsgDLC_Index).value
                #if(ws.cell(row=Row, column=MsgDLC_Index).value=='Var'):
                 #   CanMsg.DLC = '8'
                # 获取报文发送节点
                for j in range(26, MaxColumn):
                    if((ws.cell(row=Row, column=j).value)=="Tx"):
                        CanMsg.TxNode = ALLNodes[j-26]
                        break
                #print(CanMsg.ID)
                #print(CanMsg.CycleTime)
                #print(CanMsg.TxNode)
            else:#信号
                CANSig = Sig() #构造信号类对象
                CANSig.Name = ws.cell(row=Row, column=SigName_Index).value
                CANSig.SendType = ws.cell(row=Row, column=SigSendType_Index).value
                CANSig.BitLenth = ws.cell(row=Row, column=SigLenth_Index).value
                #if (CANSig.BitLenth) == 'Var':
                    #CANSig.BitLenth = '8'
                #Index = ws.cell(row=Row, column=SigBit_Index).value.find('-')
                CANSig.Bit = ws.cell(row=Row, column=SigBit_Index).value #[(Index + 1):] #' 6-5'
                CANSig.Byte = ws.cell(row=Row, column=SigByte_Index).value #[0] #'2-3'
                CANSig.Desc = ws.cell(row=Row, column=SigDesc_Index).value
                #CANSig.StartBit = (int(CANSig.Byte[0])-1)*8 + (int(CANSig.Bit[2])-1)
                CANSig.Max = ws.cell(row=Row, column=SigMax_Index).value
                CANSig.Min = ws.cell(row=Row, column=SigMin_Index).value
                CANSig.OffSet = ws.cell(row=Row, column=SigOffSet_Index).value
                CANSig.Reso = ws.cell(row=Row, column=SigReso_Index).value
                CANSig.SigValueDesc = ws.cell(row=Row, column=SigValueDesc_Index).value
                CANSig.SPN = ws.cell(row=Row, column=SigSPN_Index).value
                CANSig.Unit = ws.cell(row=Row, column=SigUnit_Index).value
                CanMsg.Signals.append(CANSig) #添加信号到Msg
                del CANSig

            if((Row<MaxRow)and(ws.cell(row=Row+1,column=MsgName_Index).value!=None)):#处理完最后一个信号,
                CanMsgs.append(CanMsg)#添加报文
                del CanMsg
            if(Row==MaxRow): #添加最后一个报文
                CanMsgs.append(CanMsg)  # 添加报文
                del CanMsg
    except:
        ExceptionInfo += "Read In Rows:%d" % (Row)
        ExceptionInfo += traceback.format_exc()
        print(ExceptionInfo)
        input('发现异常，按Eneter退出....')
    else:
        return CanMsgs

#检查Matrix的错误
FaultInfo = "Matrix Fault:\n"
def CheckFault(CanMsgs):
    global FaultInfo
    global ExceptionInfo
    MsgID = []
    i = 0
    j = 0
    #检查报文错误
    try:
        for i in range(0, len(CanMsgs)):
        #1.ID重复
            MsgID.append(CanMsgs[i].ID)
        MsgID.sort()
        for i in range(0, len(MsgID)):
            if(MsgID.count(MsgID[i])>1):
                FaultInfo += "0x%x,ID:Repeat\n"% (MsgID[i])

        #2.Msg其他错误
        for i in range(0, len(CanMsgs)):
            #1.检查ID错误
            if(CanMsgs[i].ID==None):
                FaultInfo += "Msg:%s,ID:Empty\n"%(CanMsgs[i].Name)
            #2.检查CylceTime
            if ((CanMsgs[i].SendType == 'Periodic') and (CanMsgs[i].CycleTime == None)):
                FaultInfo += "Msg:%s,CylceTime:Empty\n"%(CanMsgs[i].Name)
            #3.检查DLC
            if (CanMsgs[i].DLC == None or CanMsgs[i].DLC == 'Var'):
                FaultInfo += "Msg:%s,DLC:Empty or Var\n"%(CanMsgs[i].Name)
            #4.检查SendType
            if (CanMsgs[i].SendType == None):
                FaultInfo += "Msg:%s,SendType:Empty\n"%(CanMsgs[i].Name)
            #5.检查TxNode
            if (CanMsgs[i].TxNode == ''):
                FaultInfo += "Msg:%s,TxNode:Empty\n"%(CanMsgs[i].Name)
            #检查信号错误
            for j in range(0, len(CanMsgs[i].Signals)):
                # 1.检查SPN,可以为空
                #if(CanMsgs[i].Signals[j].SPN == None):
                 #   FaultInfo += "Signal:%s SPN:Empty\n"% (CanMsgs[i].Signals[j].Name)
                # 2.检查Byte
                if (CanMsgs[i].Signals[j].Byte == None):
                    FaultInfo += "Signal:%s,Byte:Empty\n"% (CanMsgs[i].Signals[j].Name)
                # 3.检查Bit
                if (CanMsgs[i].Signals[j].Bit == None):
                    FaultInfo += "Signal:%s,Bit:Empty\n"% (CanMsgs[i].Signals[j].Name)
                # 4.检查BitLenth
                if (CanMsgs[i].Signals[j].BitLenth == None or CanMsgs[i].Signals[j].BitLenth == 'Var'):
                    FaultInfo += "Signal:%s,BitLenth:Empty or Var\n"% (CanMsgs[i].Signals[j].Name)
                #更新Bit信息
                if(CanMsgs[i].Signals[j].Bit != None and ((CanMsgs[i].Signals[j].BitLenth != None and CanMsgs[i].Signals[j].BitLenth != 'Var'))):
                    if(len(CanMsgs[i].Signals[j].Bit)>=3):
                        CanMsgs[i].Signals[j].Bit = CanMsgs[i].Signals[j].Bit[2]
                    else:
                        CanMsgs[i].Signals[j].Bit = CanMsgs[i].Signals[j].Bit[0]

                # 5.检查Resolution
                if (CanMsgs[i].Signals[j].Reso == None):
                    FaultInfo += "Signal:%s,Reso:Empty\n"% (CanMsgs[i].Signals[j].Name)
                # 6.检查Offset
                if (CanMsgs[i].Signals[j].OffSet == None):
                    FaultInfo += "Signal:%s,OffSet:Empty\n"% (CanMsgs[i].Signals[j].Name)
                # 7.检查Min
                if (CanMsgs[i].Signals[j].Min == None):
                    FaultInfo += "Signal:%s,Min:Empty\n"% (CanMsgs[i].Signals[j].Name)
                # 8.检查Max
                if (CanMsgs[i].Signals[j].Max == None):
                    FaultInfo += "Signal:%s,Max:Empty\n"% (CanMsgs[i].Signals[j].Name)
                # 9.检查SendType
                if (CanMsgs[i].Signals[j].SendType == None):
                    FaultInfo += "Signal:%s,SendType:Empty\n"% (CanMsgs[i].Signals[j].Name)
                # 10.检查Signal Value Description
                if(CanMsgs[i].Signals[j].SigValueDesc!=None):
                    ValueList = CanMsgs[i].Signals[j].SigValueDesc.splitlines()
                    for k in range(0,len(ValueList)):
                        if(ValueList[k].find(':')==-1):
                            FaultInfo += "Signal:%s,Signal Value Description:Format Error\n" % (CanMsgs[i].Signals[j].Name)
                # 11.检查Unit
                if(CanMsgs[i].Signals[j].Unit==None):
                    FaultInfo += "Signal:%s,Unit:Empty\n" % (CanMsgs[i].Signals[j].Name)
                    CanMsgs[i].Signals[j].Unit = CanMsgs[i].Signals[j].Unit
        if(FaultInfo == "Matrix Fault:\n"):#没错误
            return 0
        else:
            return 1
        #print(FaultInfo)
    except:
        ExceptionInfo += "Check Error In Msg.Signal:%s\n" % (CanMsgs[i].Signals[j].Name)
        ExceptionInfo += traceback.format_exc()
        print(ExceptionInfo)
        WriteCheckReport()
        #sys.exit(-1)
        input('发现异常，按Eneter退出....')
#写检查报告
def WriteCheckReport():
    print("WriteCheckReport")
    filename = pwd + "\\CheckReport.csv"
    CheckReport = open(filename, mode='w+')
    CheckReport.write(FaultInfo)
    CheckReport.flush()
    CheckReport.close()

#将Matrx提取的信息写到CSV文件，用于查看
def Write2Csv():
    fileName = pwd + 'Tail.csv'
    Tail = open(fileName, mode='w+')
    All = 'MsgName,ID,DLC,TxNode,SendType,CycleTime,SigName,Desc,Byte,Bit,BitLenth,StartBit,Reso,OffSet,Min,Max,SigValueSesc\n'
    i = 0
    j = 0
    for i in range(0, len(CanMsgs)):
        All += "%s,%d,%d,%s,%s,%d\n" % (
        CanMsgs[i].Name, CanMsgs[i].ID, CanMsgs[i].DLC, CanMsgs[i].TxNode, CanMsgs[i].SendType, CanMsgs[i].CycleTime)
        for j in range(0, len(CanMsgs[i].Signals)):
            All += "%s,%s,%s,%s,%s,%d,%f,%f,%f,%f,%s\n" % (
            CanMsgs[i].Signals[j].Name, CanMsgs[i].Signals[j].Desc, CanMsgs[i].Signals[j].Byte,
            CanMsgs[i].Signals[j].Bit, CanMsgs[i].Signals[j].BitLenth, CanMsgs[i].Signals[j].StartBit,
            CanMsgs[i].Signals[j].Reso, CanMsgs[i].Signals[j].OffSet, CanMsgs[i].Signals[j].Min,
            CanMsgs[i].Signals[j].Max, CanMsgs[i].Signals[j].SigValueDesc)
    Tail.write(All)
    Tail.flush()
    Tail.close()

#写到DBC文件

def WriteToDBC(FileName):
    global ExceptionInfo
    # 0.新建DBC文件
    DBCFileName = FileName.replace('xlsx','dbc')
    DBCFile = open(DBCFileName, mode='w+')
    # 1.写基础信息,根据dbc类型进行配置
    
    # 2.写入所有的节点
    try:
        DBCFile.write('BU_: ')
        for i in range(0, len(ALLNodes)):
            DBCFile.write('%s '%ALLNodes[i])
        DBCFile.write('\n')
        # 要求写值表
        if(ValueTableFlag):
            # 3.写入值表,值表为空的不写
            # 3.1 生成值表
            for i in range(0,len(CanMsgs)):
                for j in range(0,len(CanMsgs[i].Signals)):
                    #值表非空
                    if(CanMsgs[i].Signals[j].SigValueDesc != None):
                        ValueList = CanMsgs[i].Signals[j].SigValueDesc.splitlines()
                        for k in range(0,len(ValueList)):
                            Value = int(ValueList[k].split(':')[0],16)
                            Desc = ValueList[k].split(':')[1]
                            CanMsgs[i].Signals[j].ValueTable.append([Value,Desc])
            # 3.2 写值表
            for i in range(0,len(CanMsgs)):
                for j in range(0,len(CanMsgs[i].Signals)):
                    if(CanMsgs[i].Signals[j].SigValueDesc!=None):
                        DBCFile.write('VAL_TABLE_ %s  ' % (CanMsgs[i].Signals[j].Name))
                        for n in range(0,len(CanMsgs[i].Signals[j].ValueTable)):
                            DBCFile.write('%d "%s" '%(CanMsgs[i].Signals[j].ValueTable[n][0],CanMsgs[i].Signals[j].ValueTable[n][1]))
                        DBCFile.write(' ;\n')
        # 4.写入报文和信号
        for i in range(0, len(CanMsgs)):
            DBCFile.write('\nBO_ %d %s: %s %s\n'%(CanMsgs[i].ID,CanMsgs[i].Name,CanMsgs[i].DLC,CanMsgs[i].TxNode))
            for j in range(0, len(CanMsgs[i].Signals)):
                CanMsgs[i].Signals[j].StartBit = (int(CanMsgs[i].Signals[j].Byte[0]) - 1) * 8 + (int(CanMsgs[i].Signals[j].Bit)-1)
                str1 = " SG_ %s : %d|%s@1+ (%s,%s) [%s|%s] \"%s\" Vector__XXX\n" % (CanMsgs[i].Signals[j].Name,CanMsgs[i].Signals[j].StartBit,CanMsgs[i].Signals[j].BitLenth,CanMsgs[i].Signals[j].Reso,CanMsgs[i].Signals[j].OffSet,CanMsgs[i].Signals[j].Min,CanMsgs[i].Signals[j].Max,CanMsgs[i].Signals[j].Unit)
                DBCFile.write(str1)
        # 5.写注释
        for i in range(0, len(CanMsgs)):
            DBCFile.write('\nCM_ BO_ %d "%s";\n'%(CanMsgs[i].ID,CanMsgs[i].Description))
            for j in range(0, len(CanMsgs[i].Signals)):
                DBCFile.write('CM_ SG_ %d %s "%s";\n'%(CanMsgs[i].ID,CanMsgs[i].Signals[j].Name,CanMsgs[i].Signals[j].Desc))
        # 6.写属性
        # 6.1 写基础属性
        DBCFile.write(BaseAttribute)
        #6.2 J1939属性
        if(J1939Flag):
            DBCFile.write(J1939Attribute)
        DBCFile.write(BaseAttributeDefault)
        if (J1939Flag):
            DBCFile.write(J1939AttributeDefault)
        DBCFile.write('\n')
        # 7.写报文的发送类型、周期时间、帧格式
        for i in range(0, len(CanMsgs)):
            DBCFile.write('BA_ "GenMsgSendType" BO_ %d %d;\n'%(CanMsgs[i].ID,MsgSendTypeList[CanMsgs[i].SendType]))
            #if(CanMsgs[i].SendType=='Cycle'):
            if(CanMsgs[i].SendType=='Periodic'):    
                DBCFile.write('BA_ "GenMsgCycleTime" BO_ %d %s;\n'%(CanMsgs[i].ID,CanMsgs[i].CycleTime))
            if(J1939Flag):
                DBCFile.write('BA_ "VFrameFormat" BO_ %d 3;\n'%(CanMsgs[i].ID))
            elif(Extended):
                DBCFile.write('BA_ "VFrameFormat" BO_ %d 1;\n' % (CanMsgs[i].ID))
            else:
                DBCFile.write('BA_ "VFrameFormat" BO_ %d 0;\n' % (CanMsgs[i].ID))
        # 8.写信号的SPN和发送类型
        for i in range(0, len(CanMsgs)):
            for j in range(0, len(CanMsgs[i].Signals)):
                if(J1939Flag and CanMsgs[i].Signals[j].SPN!=None):
                    DBCFile.write('BA_ "SPN" SG_ %d %s %s;\n'%(CanMsgs[i].ID,CanMsgs[i].Signals[j].Name,CanMsgs[i].Signals[j].SPN))
                DBCFile.write('BA_ "GenSigSendType" SG_ %d %s %d;\n'%(CanMsgs[i].ID,CanMsgs[i].Signals[j].Name,SigSendTypeList[CanMsgs[i].Signals[j].SendType]))
        # 9.写VAL
        for i in range(0, len(CanMsgs)):
            for j in range(0, len(CanMsgs[i].Signals)):
                if(CanMsgs[i].Signals[j].SigValueDesc != None):
                    DBCFile.write('VAL_ %d %s  '%(CanMsgs[i].ID,CanMsgs[i].Signals[j].Name))
                    for n in range(0, len(CanMsgs[i].Signals[j].ValueTable)):
                        DBCFile.write('%d "%s" ' % (CanMsgs[i].Signals[j].ValueTable[n][0], CanMsgs[i].Signals[j].ValueTable[n][1]))
                    DBCFile.write(' ;\n')
    except:
        ExceptionInfo += "GenerateDBC Error In Msg.Signal:%s\n" % (CanMsgs[i].Signals[j].Name)
        ExceptionInfo += traceback.format_exc()
        print(ExceptionInfo)
        #sys.exit(-1)
        input('发现异常，按Eneter退出....')
# 10.关闭和刷新文件
    else:
        DBCFile.flush()
        DBCFile.close()

ValueTableFlag = 1
J1939Flag = 1
Extended = 0
#配置属性
J1939Attribute = ''
J1939AttributeDefault = ''
BaseAttribute = ''
BaseAttributeDefault = ''
MsgSendTypeList = {'Periodic': 0, 'Event': 1, 'IfActive': 2, 'If Active': 2, 'PE': 3, 'CA': 4, 'OnRequest': 5, 'On Request': 5, 'Cycle': 6,
                       'On Change': 7, 'OnChange': 7, 'OnChageWithRepetition': 8}
SigSendTypeList = {'Periodic': 0, 'OnChange': 1,'On Change': 1, 'OnWrite': 2, 'On Write': 2,'IfActive': 3, 'OnChangeWithRepetition': 4,
                       'OnWriteWithRepetition': 5, 'IfActiveWithRepetition': 6, 'NoSigSendType': 7, 'Event': 8,
                       'Cycle': 9}
MaxColumn = 0
MaxRow = 0
ws = None
CanMsgs = None
def mainProcess():
    global MaxColumn, MaxRow, ws ,CanMsgs
    # 0.获取矩阵文件路径
    MatrixFileName = input("输入Excel文件路径+文件名:")
    #MatrixFileName
    # 1.加载excel文件
    ws = loadexcel(MatrixFileName)
    MaxColumn = ws.max_column
    MaxRow = ws.max_row
    # 2.读取excel文件，构造对象
    CanMsgs = GetMsg()
    #print("len(CanMsgs):%s"%len(CanMsgs))
    '''检查错误'''
    CheckResult = CheckFault(CanMsgs)
    '''写检查报告'''
    WriteCheckReport()
    '''没有错误，则写DBC文件'''
    if (CheckResult == 0):
        WriteToDBC(MatrixFileName)
        print('Generate DBC Success!')
    else:
        print('Generate DBC Fail!')

#作为单独的脚本使用时，__name__属性为__main__
if __name__=='__main__':
    mainProcess()
    input('按Eneter继续....')





